"""Низкоуровневые читатели листов Excel телеметрии ДНС-7с.

Форматы листов (по факту в «Копия ДНС-7с ДН.xlsx»):
  * Временной ряд (давления, расход, уровни): первая непустая строка кол. A —
    имя сигнала (заголовок); далее строки [datetime, текстовое число].
  * Журнал состояний («Состояние Н-X»): метаданные, затем строка-заголовок
    ['Описание тега','Значение','Время']; далее [тег, 'ВКЛ.'/'ОТКЛ.', datetime].
  * Энергия («Расход эл. энергии Н-X»): метаданные, заголовок ['Дата', '...кВт*ч','Кач.'];
    далее [строка-дата, кВт·ч, флаг качества]. Интервалы получасовые.
  * Перекачка («ДНС-7с объем перекачки.xlsx», Лист1): заголовок во 2-й строке;
    суточные [Цех, Тов.парк, ДНС, Qплан, Qзамер, Дата замера, счётчик].

Значения чисел в исходниках хранятся как ТЕКСТ ('2.740900') — парсятся в float.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


def _to_float(v) -> float:
    """Текст/число → float; иначе NaN."""
    if v is None:
        return float("nan")
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return float("nan")


def _to_datetime(v):
    """datetime/строка → pandas.Timestamp; иначе NaT."""
    if isinstance(v, datetime):
        return pd.Timestamp(v)
    if v is None:
        return pd.NaT
    return pd.to_datetime(str(v), dayfirst=True, errors="coerce")


def _iter_rows(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            yield row
    finally:
        wb.close()


def read_timeseries(path: Path, sheet: str) -> pd.DataFrame:
    """Лист «давление/расход/уровень» → DataFrame[timestamp, value, signal].

    Заголовок (имя сигнала) — первая строка, где кол. A не datetime и не пуста.
    """
    rows = []
    signal = sheet
    header_done = False
    for a, b, *_ in _iter_rows(path, sheet):
        if a is None and b is None:
            continue
        if not header_done and not isinstance(a, datetime):
            # строка-заголовок: имя сигнала в кол. A
            if a is not None and str(a).strip():
                signal = str(a).strip()
            header_done = True
            continue
        ts = _to_datetime(a)
        if ts is pd.NaT:
            continue
        rows.append((ts, _to_float(b)))
    df = pd.DataFrame(rows, columns=["timestamp", "value"])
    df["signal"] = signal
    return df


def read_journal(path: Path, sheet: str) -> pd.DataFrame:
    """Лист «Состояние Н-X» → DataFrame[timestamp, state] (state ∈ {ВКЛ, ОТКЛ})."""
    rows = []
    started = False
    for a, b, c, *_ in _iter_rows(path, sheet):
        if not started:
            if a is not None and str(a).strip().startswith("Описание тега"):
                started = True
            continue
        if a is None:
            continue
        state = (str(b).strip().rstrip(".") if b is not None else "")
        state = {"ВКЛ": "ВКЛ", "ОТКЛ": "ОТКЛ"}.get(state, state)
        ts = _to_datetime(c)
        if ts is pd.NaT:
            continue
        rows.append((ts, state))
    return pd.DataFrame(rows, columns=["timestamp", "state"])


def read_energy(path: Path, sheet: str) -> pd.DataFrame:
    """Лист «Расход эл. энергии Н-X» → DataFrame[timestamp, kwh, quality]."""
    rows = []
    started = False
    for a, b, c, *_ in _iter_rows(path, sheet):
        if not started:
            if a is not None and str(a).strip().startswith("Дата"):
                started = True
            continue
        ts = _to_datetime(a)
        if ts is pd.NaT:
            continue
        # отсечь служебные строки «Итого/Всего/Пустые»
        rows.append((ts, _to_float(b), c))
    return pd.DataFrame(rows, columns=["timestamp", "kwh", "quality"])


def read_transfer(path: Path, sheet: str = "Лист1") -> pd.DataFrame:
    """Лист объёма перекачки → DataFrame[date, plan, fact, counter]."""
    rows = []
    for row in _iter_rows(path, sheet):
        # ожидаем структуру: Цех, Тов.парк, ДНС, Qплан, Qзамер, Дата, счётчик
        if len(row) < 6:
            continue
        cex, park, dns, q_plan, q_fact, dt, *rest = row
        ts = _to_datetime(dt)
        if ts is pd.NaT:
            continue
        counter = rest[0] if rest else None
        rows.append((ts.normalize(), _to_float(q_plan), _to_float(q_fact), counter))
    return pd.DataFrame(rows, columns=["date", "plan", "fact", "counter"])
