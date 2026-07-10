"""Семантический парсер инженерных файлов «… расчет.xlsx».

Файлы — ручные расчёты энергоаудита: содержат и ВХОДЫ (паспорт насоса/ЭД, режим),
и ЭТАЛОННЫЕ ВЫХОДЫ (УРЭ, КПД, напор должный, годовые потери). Шаблоны у объектов
различаются (метки в колонке D или F; значения в разных колонках) — поэтому поиск
идёт ПО ТЕКСТУ МЕТОК, а значения берутся из колонок реальных агрегатов.

Возвращает ObjectSpec с заполненными regime и reference у каждого агрегата.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl

from ..spec import (AggregateSpec, Branch, MotorSpec, ObjectSpec, PumpSpec,
                    ReferenceOutputs, RegimeMeasurement, WaterType,
                    infer_motor_synchronous, infer_pump_kind)


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower().replace("ё", "е"))


def _num(v) -> Optional[float]:
    """Число или None (мусор #REF!/#DIV/0!/текст → None)."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".") if v is not None else ""
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


class _Sheet:
    """2D-сетка значений листа (1-индексация) с поиском по меткам."""

    def __init__(self, rows: list[tuple]):
        self.rows = rows
        self.nrow = len(rows)
        self.ncol = max((len(r) for r in rows), default=0)

    def cell(self, r: int, c: int):
        if 1 <= r <= self.nrow and 1 <= c <= len(self.rows[r - 1]):
            return self.rows[r - 1][c - 1]
        return None

    def find(self, *patterns: str, exclude: tuple[str, ...] = (),
             col: Optional[int] = None) -> Optional[tuple[int, int]]:
        """Первая ячейка, чей текст содержит любой pattern и не содержит exclude.

        Если задан col — поиск только в этой колонке (устойчивость к меткам-двойникам).
        """
        cols_range = [col] if col else None
        for r in range(1, self.nrow + 1):
            rng = cols_range or range(1, len(self.rows[r - 1]) + 1)
            for c in rng:
                t = _norm(self.cell(r, c))
                if not t:
                    continue
                if any(p in t for p in patterns) and not any(e in t for e in exclude):
                    return (r, c)
        return None

    def value_at_cols(self, r: int, cols: list[int]) -> list[Optional[float]]:
        return [_num(self.cell(r, c)) for c in cols]


def _pick_sheet(wb) -> _Sheet:
    """Выбрать лист с расчётом (содержит метку «УРЭ факт»)."""
    best = None
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        sh = _Sheet(rows)
        if sh.find("урэ факт"):
            return sh
        if best is None and sh.nrow > 5:
            best = sh
    return best


def _aggregate_columns(sh: _Sheet) -> list[int]:
    """Колонки реальных агрегатов: где и Qсут, и рвых, и рвх — положительные числа."""
    def numeric_cols(anchor: tuple[int, int] | None) -> set[int]:
        if not anchor:
            return set()
        r = anchor[0]
        return {c for c in range(anchor[1] + 1, len(sh.rows[r - 1]) + 1)
                if (_num(sh.cell(r, c)) or 0) > 0}

    qsut = sh.find("суточная перекачка")
    pout = sh.find("давление на выходе", "давление на вых")
    pin = sh.find("давление на входе", "давление на вх")
    cols = numeric_cols(qsut) & numeric_cols(pout) & numeric_cols(pin)
    # отсечь явный мусор: оставить не более 4 колонок подряд
    return sorted(cols)[:4]


def _passport_block(sh: _Sheet, header_patterns: tuple[str, ...],
                    field_map: dict[str, tuple[str, ...]], n: int) -> list[dict]:
    """Распарсить блок паспорта (насоса/ЭД): по заголовку → карта колонок → n строк."""
    hdr = sh.find(*header_patterns)
    if not hdr:
        return [{} for _ in range(n)]
    r0, c0 = hdr
    # карта поле→колонка по заголовочной строке
    colmap: dict[str, int] = {}
    name_col = c0  # столбец с типом/моделью
    for c in range(c0, len(sh.rows[r0 - 1]) + 1):
        t = _norm(sh.cell(r0, c))
        for field, pats in field_map.items():
            if field not in colmap and any(p in t for p in pats):
                colmap[field] = c
    # строки агрегатов идут ниже заголовка; берём те, где есть число в любой числовой колонке
    out: list[dict] = []
    r = r0 + 1
    while r <= sh.nrow and len(out) < n:
        has_num = any(_num(sh.cell(r, c)) is not None for c in colmap.values())
        model = sh.cell(r, name_col)
        if has_num and model is not None:
            rec = {"model": str(model).strip()}
            for field, c in colmap.items():
                rec[field] = _num(sh.cell(r, c))
            out.append(rec)
        elif out:        # блок закончился
            break
        r += 1
    while len(out) < n:
        out.append({})
    return out


# Канонические метки строк режима/результатов: поле → (паттерны, исключения).
# Значение читается на колонках агрегатов.
_REGIME_LABELS = {
    "t": (("время работы на",), ()),
    "q_day": (("суточная перекачка",), ()),
    "w": (("суточный расход ээ", "расход эл. энергии", "расход ээ"), ()),
    "p_electric": (("электрическая мощность",), ()),
    "rho": (("плотность",), ()),
    "p_in": (("давление на входе", "давление на вх"), ()),
    "p_out": (("давление на выходе", "давление на вых"), ()),
    "q_fact": (("фактич произв",), ()),
    "p_bg": (("давление на бг",), ()),
    "t_year": (("годовая нараб",), ()),
}
_REF_LABELS = {
    "h_fact": (("напор",), ("должн", "снижен", "на в сут")),
    "h_due": (("напор должн",), ()),
    "p_hydraulic": (("гидравлическая мощность",), ("на бг",)),
    "eta_fact": (("кпд на среднесут",), ()),
    "eta_nom": (("кпд на в ном",), ()),
    "sec_fact": (("урэ факт",), ()),
    "sec_calc": (("урэ расч",), ()),
    "load_factor": (("коэф загруз",), ()),
    "dw_efficiency": (("wкпд",), ()),
    "dw_throttle": (("wдрос", "wдросс"), ()),
}


def _regime_label_col(sh: _Sheet) -> Optional[int]:
    """Колонка меток режима = колонка метки «Суточная перекачка»."""
    anchor = sh.find("суточная перекачка")
    return anchor[1] if anchor else None


def _read_label_values(sh: _Sheet, labels: dict, cols: list[int],
                       label_col: Optional[int]) -> dict[str, list]:
    out: dict[str, list] = {}
    for field, (pats, excl) in labels.items():
        anchor = sh.find(*pats, exclude=excl, col=label_col)
        if anchor is None:                       # запас: глобальный поиск
            anchor = sh.find(*pats, exclude=excl)
        out[field] = sh.value_at_cols(anchor[0], cols) if anchor else [None] * len(cols)
    return out


_UNIT_SCALE = {"тыс. квт": 1000.0, "тыс квт": 1000.0}


def _dw_scale(sh: _Sheet, label_pats: tuple[str, ...]) -> float:
    """Множитель единиц для годовых потерь (тыс. кВт·ч → ×1000)."""
    anchor = sh.find(*label_pats)
    if not anchor:
        return 1.0
    r, c = anchor
    for cc in range(c, min(c + 4, len(sh.rows[r - 1]) + 1)):
        t = _norm(sh.cell(r, cc))
        for key, scale in _UNIT_SCALE.items():
            if key in t:
                return scale
    return 1.0


_WATER_BY_PATH = {"пресная": WaterType.fresh, "агрессив": WaterType.aggressive,
                  "пластов": WaterType.formation}


def _water_type(path: Path) -> WaterType:
    p = _norm(str(path))
    for key, wt in _WATER_BY_PATH.items():
        if key in p:
            return wt
    return WaterType.fresh


def parse_calc_file(path: Path, object_id: str, object_name: str) -> ObjectSpec:
    """Распарсить «… расчет.xlsx» в ObjectSpec (входы + эталонные выходы)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sh = _pick_sheet(wb)
    finally:
        wb.close()
    if sh is None:
        raise ValueError(f"не найден лист расчёта в {path}")

    cols = _aggregate_columns(sh)
    if not cols:
        raise ValueError(f"не определены колонки агрегатов в {path}")

    pumps = _passport_block(sh, ("тип насоса",), {
        "h_nom": ("нном",), "q_nom": ("qном",), "power_nom": ("pпотреб", "потреб"),
        "n_rpm": ("об/мин",), "eta_nom": ("кпд",), "p_motor_rec": ("рекоменд",)}, len(cols))
    # ограничить число агрегатов количеством паспортных строк насосов (с моделью)
    n_real = sum(1 for p in pumps if p.get("model"))
    if n_real:
        cols = cols[:n_real]
        pumps = pumps[:len(cols)]
    motors = _passport_block(sh, ("тип эл",), {
        "voltage_kv": ("u, кв", "u,кв"), "i_nom": ("iном",), "p_nom": ("pном",),
        "n_rpm": ("об/мин",), "eta_nom": ("кпд",), "cos_phi": ("cos",)}, len(cols))

    label_col = _regime_label_col(sh)
    regime = _read_label_values(sh, _REGIME_LABELS, cols, label_col)
    ref = _read_label_values(sh, _REF_LABELS, cols, label_col)
    dw_scale = _dw_scale(sh, ("wкпд",))

    water = _water_type(path)
    aggregates: list[AggregateSpec] = []
    for i, col in enumerate(cols):
        p = pumps[i] if i < len(pumps) else {}
        m = motors[i] if i < len(motors) else {}
        pump_model = p.get("model", "")
        pump = PumpSpec(
            model=pump_model, kind=infer_pump_kind(pump_model, p.get("n_rpm")),
            q_nom=p.get("q_nom"), h_nom=p.get("h_nom"), eta_nom=p.get("eta_nom"),
            power_nom=p.get("power_nom"), n_rpm=p.get("n_rpm"))
        motor = MotorSpec(
            model=m.get("model", ""), synchronous=infer_motor_synchronous(m.get("model", "")),
            p_nom=m.get("p_nom"), eta_nom=m.get("eta_nom"),
            cos_phi=m.get("cos_phi"), voltage_kv=m.get("voltage_kv"), i_nom=m.get("i_nom"),
            n_rpm=m.get("n_rpm"))

        rho = regime["rho"][i] or 1000.0
        p_in, p_out = regime["p_in"][i], regime["p_out"][i]
        if p_in is None or p_out is None:
            continue
        # Правдоподобие: давления ППД — единицы МПа (макс ~16 по ограничениям).
        # Значения в сотни/тысячи — захват чужой колонки (напор в м, проектные
        # значения) → колонка не является реальным агрегатом, пропускаем.
        if p_out > 30.0 or p_in > 30.0:
            continue
        # Ноль в ячейках расхода/энергии/наработки — «не заполнено» (черновики
        # инженеров, напр. W=0 у КНС-85 НА-1), а не физический ноль → None,
        # чтобы ядро выбрало запасной путь (УРЭ_ф = P_эл/Q вместо W/Q_сут).
        _pos = lambda v: v if (v is not None and v > 0) else None  # noqa: E731
        rm = RegimeMeasurement(
            rho=rho, p_in=p_in, p_out=p_out,
            q_day=_pos(regime["q_day"][i]), t=_pos(regime["t"][i]),
            w=_pos(regime["w"][i]), p_electric=_pos(regime["p_electric"][i]),
            q_fact=_pos(regime["q_fact"][i]), p_bg=_pos(regime["p_bg"][i]),
            t_year=_pos(regime["t_year"][i]))
        ro = ReferenceOutputs(
            h_fact=ref["h_fact"][i], h_due=ref["h_due"][i],
            eta_fact=ref["eta_fact"][i], eta_nom=ref["eta_nom"][i],
            sec_fact=ref["sec_fact"][i], sec_calc=ref["sec_calc"][i],
            load_factor=ref["load_factor"][i], p_hydraulic=ref["p_hydraulic"][i],
            dw_efficiency=(ref["dw_efficiency"][i] * dw_scale) if ref["dw_efficiency"][i] is not None else None,
            dw_throttle=(ref["dw_throttle"][i] * dw_scale) if ref["dw_throttle"][i] is not None else None,
            t_year=regime["t_year"][i])
        aggregates.append(AggregateSpec(
            id=f"НА-{i + 1}", role="работа", pump=pump, motor=motor,
            regime=rm, reference=ro))

    return ObjectSpec(
        id=object_id, name=object_name, water_type=water, branch=Branch.kns,
        source=str(path), aggregates=aggregates)
